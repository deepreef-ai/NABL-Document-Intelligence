#!/usr/bin/env python3
"""Build an Excel workbook that shows, field by field, what the ground truth
says, what the pipeline predicted, and how each pair was scored — then the
accuracy/precision/recall those rows add up to.

The point of this workbook over a summary table is auditability: every number
on the Metrics sheet is the sum of rows on the comparison sheets, so a
disagreement can be inspected rather than argued about.

MATCHING
Field names are compared exactly. Values are compared on normalised text.
"Renamed" detection — the same value present under a different key — uses
TOKEN-BOUNDARY containment, not raw character containment: character
containment credits "NA" as finding "...in a client packaging..." because the
letters n,a occur inside it, which inflated an earlier version of this report
by about 7 points.

VERDICTS (one row per ground-truth field, plus one per unmatched prediction)
  TP              name matched and value matched — the only correct outcome
                  under strict field+value scoring
  WRONG VALUE     name matched, value differed
  FN (renamed)    the ground-truth value IS in the output, under another key
  FN (missed)     the value is not in the output at all
  FP (extra)      predicted field with no ground-truth counterpart

Usage:
    python scripts/build_comparison_workbook.py
    python scripts/build_comparison_workbook.py --predictions <dir> --out cmp.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

DEFAULT_ROOT = r"G:\Shared drives\Product & Engineering\Projects\NABL Document Intelligence"

MAX_CELL = 400  # keep cells readable; the full value stays in the prediction JSON


def norm(v) -> str:
    text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, sort_keys=True)
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def toks(v) -> str:
    """Space-separated tokens, decimals kept intact, so containment can be
    tested on word boundaries (see MATCHING in the module docstring)."""
    text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, sort_keys=True)
    t = re.sub(r"[^a-z0-9.]+", " ", str(text).lower())
    t = re.sub(r"(?<!\d)\.|\.(?!\d)", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def contained(a: str, b: str) -> bool:
    if not a or not b:
        return False
    return a == b or f" {a} " in f" {b} " or f" {b} " in f" {a} "


def cell(v) -> str:
    if v is None:
        return ""
    text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False)
    text = str(text)
    return text if len(text) <= MAX_CELL else text[:MAX_CELL] + " ... (+%d chars)" % (len(text) - MAX_CELL)


def build_rows(labelled_dir: Path, predictions_dir: Path):
    field_rows: list[dict] = []
    test_rows: list[dict] = []
    doc_rows: list[dict] = []

    for label_path in sorted(labelled_dir.glob("*/*.json")):
        stem = label_path.stem
        gt = json.loads(label_path.read_text(encoding="utf-8"))
        pred_path = predictions_dir / (stem + ".json")
        pred = json.loads(pred_path.read_text(encoding="utf-8")) if pred_path.exists() else {}
        err = pred.get("pipeline_error")

        g = {k: v for k, v in (gt.get("fields") or {}).items() if v is not None}
        p = {k: v for k, v in (pred.get("fields") or {}).items() if v is not None}
        p_tok = {k: toks(v) for k, v in p.items()}

        counts = dict(tp=0, wrong=0, renamed=0, missed=0, extra=0)
        claimed: set[str] = set()   # predicted keys already explained by a ground-truth row

        for k in sorted(g):
            gv = g[k]
            if k in p:
                claimed.add(k)
                if norm(p[k]) == norm(gv):
                    verdict, why = "TP", "name and value both match"
                    counts["tp"] += 1
                else:
                    verdict, why = "WRONG VALUE", "name matches, value differs"
                    counts["wrong"] += 1
                field_rows.append(dict(
                    document=stem, gt_field=k, gt_value=cell(gv),
                    pred_field=k, pred_value=cell(p[k]),
                    verdict=verdict, note=why, found_under=""))
                continue

            # Prefer the predicted key whose value matches EXACTLY over one
            # that merely contains it. Taking the first partial match in key
            # order attributed 003_Lab-report's work_sht_dttm
            # ("17/02/2016 07:34:27PM") to `date` ("17/02/2016") while the
            # identical full value sat under another key, which then showed up
            # as an unrelated "extra" row — confusing to audit even though the
            # TP/FP/FN totals are unaffected.
            gt_tok, gt_norm = toks(gv), norm(gv)
            hit = next((pk for pk in sorted(p) if norm(p[pk]) == gt_norm), None)
            if hit is None:
                hit = next((pk for pk in sorted(p) if contained(gt_tok, p_tok[pk])), None)
            if hit is not None:
                claimed.add(hit)
                counts["renamed"] += 1
                field_rows.append(dict(
                    document=stem, gt_field=k, gt_value=cell(gv),
                    pred_field="", pred_value=cell(p[hit]),
                    verdict="FN (renamed)",
                    note="value is present but under a different key",
                    found_under=hit))
            else:
                counts["missed"] += 1
                field_rows.append(dict(
                    document=stem, gt_field=k, gt_value=cell(gv),
                    pred_field="", pred_value="",
                    verdict="FN (missed)",
                    note=(err[:120] if err else "value not found anywhere in the output"),
                    found_under=""))

        for k in sorted(set(p) - set(g) - claimed):
            counts["extra"] += 1
            field_rows.append(dict(
                document=stem, gt_field="", gt_value="",
                pred_field=k, pred_value=cell(p[k]),
                verdict="FP (extra)",
                note="predicted field has no ground-truth counterpart",
                found_under=""))

        # ---- test-table rows, matched on test name
        g_tests = {norm(r.get("test_name")): r for r in (gt.get("tests") or []) if r.get("test_name")}
        p_tests = {norm(r.get("test_name")): r for r in (pred.get("tests") or []) if r.get("test_name")}
        t_counts = dict(tp=0, wrong=0, missed=0, extra=0)

        for key in sorted(g_tests):
            gr = g_tests[key]
            if key in p_tests:
                pr = p_tests[key]
                ok = norm(gr.get("result")) == norm(pr.get("result"))
                t_counts["tp" if ok else "wrong"] += 1
                test_rows.append(dict(
                    document=stem, test_name=cell(gr.get("test_name")),
                    gt_result=cell(gr.get("result")), pred_result=cell(pr.get("result")),
                    gt_unit=cell(gr.get("unit")), pred_unit=cell(pr.get("unit")),
                    gt_range=cell(gr.get("reference_range")), pred_range=cell(pr.get("reference_range")),
                    verdict="TP" if ok else "WRONG VALUE",
                    unit_ok="yes" if norm(gr.get("unit")) == norm(pr.get("unit")) else "no",
                    range_ok="yes" if norm(gr.get("reference_range")) == norm(pr.get("reference_range")) else "no"))
            else:
                t_counts["missed"] += 1
                test_rows.append(dict(
                    document=stem, test_name=cell(gr.get("test_name")),
                    gt_result=cell(gr.get("result")), pred_result="",
                    gt_unit=cell(gr.get("unit")), pred_unit="",
                    gt_range=cell(gr.get("reference_range")), pred_range="",
                    verdict="FN (missed)", unit_ok="", range_ok=""))

        for key in sorted(set(p_tests) - set(g_tests)):
            pr = p_tests[key]
            t_counts["extra"] += 1
            test_rows.append(dict(
                document=stem, test_name=cell(pr.get("test_name")),
                gt_result="", pred_result=cell(pr.get("result")),
                gt_unit="", pred_unit=cell(pr.get("unit")),
                gt_range="", pred_range=cell(pr.get("reference_range")),
                verdict="FP (extra)", unit_ok="", range_ok=""))

        tp = counts["tp"]
        fp = counts["wrong"] + counts["extra"]
        fn = counts["wrong"] + counts["renamed"] + counts["missed"]
        doc_rows.append(dict(
            document=stem, status="failed" if err else "ok", pipeline_error=(err or "")[:150],
            doc_type=pred.get("doc_type") or "", pages=pred.get("page_count") or "",
            extraction_source=pred.get("extraction_source") or "",
            gt_fields=len(g), pred_fields=len(p),
            tp=tp, wrong_value=counts["wrong"], fn_renamed=counts["renamed"],
            fn_missed=counts["missed"], fp_extra=counts["extra"],
            precision=(round(tp / (tp + fp), 4) if tp + fp else None),
            recall=(round(tp / (tp + fn), 4) if tp + fn else None),
            gt_test_rows=len(g_tests), test_tp=t_counts["tp"],
            test_wrong=t_counts["wrong"], test_missed=t_counts["missed"],
            test_extra=t_counts["extra"]))

    return field_rows, test_rows, doc_rows


def _prf(tp: int, fp: int, fn: int):
    p = tp / (tp + fp) if tp + fp else 0.0
    r = tp / (tp + fn) if tp + fn else 0.0
    f = 2 * p * r / (p + r) if p + r else 0.0
    return p, r, f


def build_workbook(field_rows, test_rows, doc_rows, out_path: Path, source: Path) -> dict:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1B5FD1")
    bold = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    FILLS = {
        "TP": PatternFill("solid", fgColor="D6F0DE"),
        "WRONG VALUE": PatternFill("solid", fgColor="FFF2CC"),
        "FN (renamed)": PatternFill("solid", fgColor="DDE8FB"),
        "FN (missed)": PatternFill("solid", fgColor="FBD9D9"),
        "FP (extra)": PatternFill("solid", fgColor="F1E2F7"),
    }

    wb = Workbook()

    def sheet(title, rows, widths, pct_cols=()):
        ws = wb.create_sheet(title)
        if not rows:
            ws.append(["(no rows)"])
            return ws
        cols = list(rows[0].keys())
        ws.append([c.replace("_", " ") for c in cols])
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        verdict_col = cols.index("verdict") + 1 if "verdict" in cols else None
        for r in rows:
            ws.append([r.get(c) for c in cols])
            if verdict_col:
                fill = FILLS.get(str(r.get("verdict")))
                if fill:
                    ws.cell(row=ws.max_row, column=verdict_col).fill = fill
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap_top
        for name in pct_cols:
            if name in cols:
                idx = cols.index(name) + 1
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = "0.0%"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ---- pooled counts, straight off the rows
    def v(name):
        return sum(1 for r in field_rows if r["verdict"] == name)

    TP, WRONG, RENAMED, MISSED, EXTRA = v("TP"), v("WRONG VALUE"), v("FN (renamed)"), v("FN (missed)"), v("FP (extra)")

    # FP/FN are derived from the TOTALS, not from the verdict buckets.
    # Deriving FP as WRONG+EXTRA would quietly exempt the predicted fields
    # that appear only as a "renamed" target — a field holding a real value
    # under the wrong key is still not the pair that was asked for, and
    # forgiving it here while the sheet says "strict" overstated precision by
    # about 8 points in an earlier version of this script. Under strict
    # field-value scoring every predicted field that is not an exact TP is an
    # FP, and every ground-truth field that is not an exact TP is an FN; the
    # verdict buckets stay as the diagnostic that says WHERE those errors are.
    GT_TOTAL = sum(d["gt_fields"] for d in doc_rows)
    PRED_TOTAL = sum(d["pred_fields"] for d in doc_rows)
    FP = PRED_TOTAL - TP
    FN = GT_TOTAL - TP

    def tv(name):
        return sum(1 for r in test_rows if r["verdict"] == name)

    tTP, tWRONG, tMISSED, tEXTRA = tv("TP"), tv("WRONG VALUE"), tv("FN (missed)"), tv("FP (extra)")
    # Test rows are keyed by normalised test name, so a document repeating the
    # same analyte name collapses to one row on each side; these totals are
    # therefore the DEDUPLICATED counts, which is what the matching used.
    tGT_TOTAL = tTP + tWRONG + tMISSED
    tPRED_TOTAL = tTP + tWRONG + tEXTRA
    tFP = tPRED_TOTAL - tTP
    tFN = tGT_TOTAL - tTP

    P, R, F1 = _prf(TP, FP, FN)
    tP, tR, tF1 = _prf(tTP, tFP, tFN)
    aTP, aFP, aFN = TP + tTP, FP + tFP, FN + tFN
    aP, aR, aF1 = _prf(aTP, aFP, aFN)

    ws = wb.active
    ws.title = "Metrics"
    ws.append(["Accuracy, precision and recall"])
    ws["A1"].font = Font(bold=True, size=14, color="1B5FD1")
    ws.append(["Source predictions: " + str(source)])
    ws.append(["Documents: %d   |   ground-truth fields: %d   |   predicted fields: %d   |   ground-truth test rows: %d"
               % (len(doc_rows), GT_TOTAL, PRED_TOTAL, tGT_TOTAL)])
    ws.append(["Every number below is the sum of rows on the Field comparison / Test rows sheets — filter the "
               "'verdict' column there to see exactly which rows produced it."])
    ws.append([])

    ws.append(["What counts as correct"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["A prediction is CORRECT only when the field name AND the value both match (strict field-value pair "
               "scoring). A right name with a wrong value counts as an error on both sides, because a wrong value "
               "in a correctly-named slot is not a usable answer."])
    ws.append([])

    ws.append(["Verdict", "Count", "Counts toward", "Meaning"])
    for c in ws[ws.max_row]:
        c.font = head_font
        c.fill = head_fill
    for label, n, toward, meaning in [
        ("TP", TP, "TP", "Name matched and value matched — the only correct outcome."),
        ("WRONG VALUE", WRONG, "FP + FN", "Name matched, value differed."),
        ("FN (renamed)", RENAMED, "FP + FN", "The value IS in the output, but under a different key — so the "
                                             "ground-truth pair is missing AND the predicted pair is not one "
                                             "that was asked for."),
        ("FN (missed)", MISSED, "FN", "The value is not in the output at all."),
        ("FP (extra)", EXTRA, "FP", "Predicted field with no ground-truth counterpart."),
    ]:
        ws.append([label, n, toward, meaning])
        ws.cell(row=ws.max_row, column=1).fill = FILLS[label]
    ws.append([])
    ws.append(["How FP and FN are counted"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["FP = every predicted field that is not an exact TP  =  %d predicted - %d TP  =  %d\n"
               "FN = every ground-truth field that is not an exact TP  =  %d ground truth - %d TP  =  %d\n"
               "The verdict buckets above explain WHERE those errors come from; they are not separate "
               "exemptions. A field holding the right value under the wrong key is still not the pair that "
               "was asked for, so it is charged on both sides."
               % (PRED_TOTAL, TP, FP, GT_TOTAL, TP, FN)])
    ws.row_dimensions[ws.max_row].height = 64
    ws.append([])

    ws.append(["Measure", "Formula", "With these numbers", "Result"])
    for c in ws[ws.max_row]:
        c.font = head_font
        c.fill = head_fill

    for title, tp, fp, fn, p, r, f in [
        ("DOCUMENT FIELDS", TP, FP, FN, P, R, F1),
        ("TEST-TABLE ROWS", tTP, tFP, tFN, tP, tR, tF1),
        ("ALL DATA COMBINED", aTP, aFP, aFN, aP, aR, aF1),
    ]:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.append(["Precision", "TP / (TP + FP)", "%d / (%d + %d) = %d" % (tp, tp, fp, tp + fp), p])
        ws.append(["Recall", "TP / (TP + FN)", "%d / (%d + %d) = %d" % (tp, tp, fn, tp + fn), r])
        ws.append(["F1", "2 x P x R / (P + R)", "the two above, balanced", f])
        ws.append(["Accuracy (strict)", "TP / (TP + FP + FN)", "%d / %d" % (tp, tp + fp + fn),
                   (tp / (tp + fp + fn) if tp + fp + fn else 0.0)])
        for row in range(ws.max_row - 3, ws.max_row + 1):
            ws.cell(row=row, column=4).number_format = "0.0%"
            ws.cell(row=row, column=4).font = bold
        ws.append([])

    ws.append(["Note on 'Accuracy'"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["True negatives do not exist here: field names are invented from each document's own labels, so "
               "there is no fixed list of fields that were correctly NOT produced. Textbook accuracy "
               "(TP+TN)/total is therefore not computable. The accuracy above is the Jaccard form, "
               "TP/(TP+FP+FN) — of every field either side put forward, the share both agreed on. It is the "
               "strictest honest single number for this task."])

    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).alignment = wrap_top
    for col, w in zip("ABCD", (26, 24, 30, 66)):
        ws.column_dimensions[col].width = w
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    ws.row_dimensions[4].height = 28

    sheet("Field comparison", field_rows, (34, 30, 56, 30, 56, 16, 44, 26))
    sheet("Test rows", test_rows, (34, 34, 18, 18, 14, 14, 20, 20, 15, 9, 9))
    sheet("Per-document", doc_rows,
          (34, 9, 34, 28, 7, 20, 10, 11, 7, 12, 12, 11, 10, 11, 10, 12, 9, 12, 11),
          pct_cols=("precision", "recall"))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "fields": dict(tp=TP, fp=FP, fn=FN, p=P, r=R, f1=F1),
        "tests": dict(tp=tTP, fp=tFP, fn=tFN, p=tP, r=tR, f1=tF1),
        "combined": dict(tp=aTP, fp=aFP, fn=aFN, p=aP, r=aR, f1=aF1),
        "breakdown": dict(tp=TP, wrong=WRONG, renamed=RENAMED, missed=MISSED, extra=EXTRA),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--labelled", default=os.path.join(DEFAULT_ROOT, "labelled_dataset"))
    ap.add_argument("--predictions", default=os.path.join(DEFAULT_ROOT, "predictions", "predictions"))
    ap.add_argument("--out", default=os.path.join(DEFAULT_ROOT, "ground_truth_comparison.xlsx"))
    args = ap.parse_args()

    labelled_dir, pred_dir, out = Path(args.labelled), Path(args.predictions), Path(args.out)
    if not labelled_dir.is_dir():
        print("labelled dataset not found: %s" % labelled_dir, file=sys.stderr)
        return 1
    if not pred_dir.is_dir():
        print("predictions not found: %s" % pred_dir, file=sys.stderr)
        return 1

    field_rows, test_rows, doc_rows = build_rows(labelled_dir, pred_dir)
    m = build_workbook(field_rows, test_rows, doc_rows, out, pred_dir)

    print("Wrote %s" % out)
    print("  documents %d | field rows %d | test rows %d" % (len(doc_rows), len(field_rows), len(test_rows)))
    b = m["breakdown"]
    print("  verdicts: TP=%d  WRONG VALUE=%d  FN renamed=%d  FN missed=%d  FP extra=%d"
          % (b["tp"], b["wrong"], b["renamed"], b["missed"], b["extra"]))
    for name in ("fields", "tests", "combined"):
        d = m[name]
        print("  %-9s TP=%5d FP=%5d FN=%5d | P=%5.1f%% R=%5.1f%% F1=%5.1f%%"
              % (name, d["tp"], d["fp"], d["fn"], d["p"] * 100, d["r"] * 100, d["f1"] * 100))
    return 0


if __name__ == "__main__":
    sys.exit(main())
