#!/usr/bin/env python3
"""Build an Excel report comparing two benchmark runs field-by-field:
confusion-matrix counts, precision, recall, F1 and accuracy, for the CURRENT
run and any PREVIOUS run, side by side.

Both runs are read from their cached predictions/ folders, so this makes no
LLM calls and can be re-run freely.

A note on the confusion matrix, because it is not the textbook 2x2 here.
This is open-set extraction: there is no fixed list of possible field names
to answer "which fields correctly did NOT appear", so TRUE NEGATIVES are
undefined. TP/FP/FN are all well defined and are what precision/recall/F1
need; the sheet states TN as "n/a (open set)" rather than inventing a zero
that would make textbook accuracy = (TP+TN)/total look artificially precise.
The accuracy reported instead is coverage: of every datum the ground truth
records, how much did the pipeline extract — which is the question a reader
actually means by "accuracy" for an extraction task.

Two levels are reported because they answer different questions:
  KEY level   — did it use the same field NAME as the human? (strict; this is
                what matters for filling a named form slot)
  VALUE level — did it extract the information at all, under any name?
                (naming-independent; this is reading ability)

Usage:
    python scripts/build_benchmark_report.py
    python scripts/build_benchmark_report.py --previous D:\\pip_tmp\\baseline_preds
    python scripts/build_benchmark_report.py --out report.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # backend/ on sys.path

DEFAULT_ROOT = r"G:\Shared drives\Product & Engineering\Projects\NABL Document Intelligence"


def _norm_value(v) -> str:
    text = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, sort_keys=True)
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def _load_predictions(predictions_dir: Path) -> dict[str, dict]:
    out = {}
    if predictions_dir.is_dir():
        for p in sorted(predictions_dir.glob("*.json")):
            out[p.stem] = json.loads(p.read_text(encoding="utf-8"))
    return out


def evaluate(labelled_dir: Path, predictions_dir: Path) -> dict:
    """Per-document and pooled counts for one run."""
    predictions = _load_predictions(predictions_dir)
    rows, totals = [], {
        "gt_fields": 0, "pred_fields": 0,
        "key_tp": 0, "key_fp": 0, "key_fn": 0,
        "value_correct": 0, "value_wrong": 0,
        "gt_values": 0, "gt_values_found": 0,
        "gt_tests": 0, "gt_tests_found": 0,
        "documents": 0, "failed": 0,
    }

    for label_path in sorted(labelled_dir.glob("*/*.json")):
        stem = label_path.stem
        gt = json.loads(label_path.read_text(encoding="utf-8"))
        pred = predictions.get(stem, {"fields": {}, "pipeline_error": "no prediction"})

        gt_fields = {k: v for k, v in (gt.get("fields") or {}).items() if v is not None}
        pred_fields = {k: v for k, v in (pred.get("fields") or {}).items() if v is not None}

        gt_keys, pred_keys = set(gt_fields), set(pred_fields)
        tp_keys = gt_keys & pred_keys
        key_tp, key_fp, key_fn = len(tp_keys), len(pred_keys - gt_keys), len(gt_keys - pred_keys)

        # Of the keys BOTH sides have, how many values actually agree.
        value_correct = sum(1 for k in tp_keys if _norm_value(gt_fields[k]) == _norm_value(pred_fields[k]))
        value_wrong = len(tp_keys) - value_correct

        # Naming-independent: is the ground-truth value present under ANY key.
        pred_vals = [_norm_value(v) for v in pred_fields.values() if _norm_value(v)]
        gt_vals = [_norm_value(v) for v in gt_fields.values() if _norm_value(v)]
        gt_found = sum(1 for g in gt_vals if any(g in p or p in g for p in pred_vals))

        test_rows = [r for r in (gt.get("tests") or []) if r.get("result") is not None and _norm_value(r["result"])]
        tests_found = sum(
            1 for r in test_rows if any(_norm_value(r["result"]) in p or p in _norm_value(r["result"]) for p in pred_vals)
        )

        failed = pred.get("pipeline_error") is not None
        rows.append({
            "document": stem,
            "status": "failed" if failed else "ok",
            "gt_fields": len(gt_fields), "pred_fields": len(pred_fields),
            "key_tp": key_tp, "key_fp": key_fp, "key_fn": key_fn,
            "value_correct": value_correct, "value_wrong": value_wrong,
            "gt_values": len(gt_vals), "gt_values_found": gt_found,
            "gt_tests": len(test_rows), "gt_tests_found": tests_found,
        })

        totals["documents"] += 1
        totals["failed"] += int(failed)
        for k in ("gt_fields", "pred_fields", "key_tp", "key_fp", "key_fn", "value_correct", "value_wrong",
                  "gt_values", "gt_values_found", "gt_tests", "gt_tests_found"):
            totals[k] += rows[-1][k]

    return {"rows": rows, "totals": totals, "metrics": _metrics(totals)}


def _metrics(t: dict) -> dict:
    def ratio(n, d):
        return round(n / d, 4) if d else None

    p = ratio(t["key_tp"], t["key_tp"] + t["key_fp"])
    r = ratio(t["key_tp"], t["key_tp"] + t["key_fn"])
    f1 = round(2 * p * r / (p + r), 4) if p and r and (p + r) else (0.0 if p is not None and r is not None else None)
    matched = t["value_correct"] + t["value_wrong"]

    # STRICT field-value pair scoring: a prediction is correct only when the
    # field NAME and the VALUE both match. This is entity-level scoring, the
    # standard for open-set key-value extraction, and it is the metric that
    # matches "extract the field and its value" as a product goal: a right
    # name with a wrong value is not a usable answer, so it counts as an error
    # on both sides rather than as a hit.
    pair_tp = t["value_correct"]
    pair_fp = t["pred_fields"] - pair_tp   # every produced pair that is not exactly right
    pair_fn = t["gt_fields"] - pair_tp     # every ground-truth pair not reproduced exactly
    sp = ratio(pair_tp, pair_tp + pair_fp)
    sr = ratio(pair_tp, pair_tp + pair_fn)
    sf1 = round(2 * sp * sr / (sp + sr), 4) if sp and sr and (sp + sr) else (
        0.0 if sp is not None and sr is not None else None)
    overall_total = t["gt_values"] + t["gt_tests"]
    overall_hit = t["gt_values_found"] + t["gt_tests_found"]
    return {
        "strict_pair_precision": sp,
        "strict_pair_recall": sr,
        "strict_pair_f1": sf1,
        "strict_pair_tp": pair_tp,
        "strict_pair_fp": pair_fp,
        "strict_pair_fn": pair_fn,
        "value_wrong_on_matched_keys": t["value_wrong"],
        "key_precision": p,
        "key_recall": r,
        "key_f1": f1,
        "value_accuracy_on_matched_keys": ratio(t["value_correct"], matched),
        "field_coverage": ratio(t["gt_values_found"], t["gt_values"]),
        "test_result_coverage": ratio(t["gt_tests_found"], t["gt_tests"]),
        "overall_accuracy": ratio(overall_hit, overall_total),
        "documents": t["documents"],
        "documents_failed": t["failed"],
    }


def _pct(v) -> str:
    return "n/a" if v is None else f"{v:.1%}"


def pick_examples(labelled_dir: Path, predictions_dir: Path) -> dict:
    """One REAL example of each bucket, taken from the actual data rather
    than invented — a confusion matrix only clicks when you can see which of
    your own fields landed in which box."""
    predictions = _load_predictions(predictions_dir)
    out = {"tp": None, "fp": None, "fn": None}
    for label_path in sorted(labelled_dir.glob("*/*.json")):
        stem = label_path.stem
        gt = json.loads(label_path.read_text(encoding="utf-8"))
        pred = predictions.get(stem) or {}
        g = {k for k, v in (gt.get("fields") or {}).items() if v is not None}
        p = {k for k, v in (pred.get("fields") or {}).items() if v is not None}
        for bucket, keys in (("tp", sorted(g & p)), ("fp", sorted(p - g)), ("fn", sorted(g - p))):
            if out[bucket] is None and keys:
                out[bucket] = (stem, keys[0])
        if all(out.values()):
            break
    return out


def _confusion_sheet(wb, current: dict, previous: dict | None, examples: dict) -> None:
    """A confusion matrix sheet written for someone meeting one for the first
    time: what each box means, one of their own fields in each box, the grid
    itself, then the arithmetic spelled out with the real numbers
    substituted, so no formula has to be taken on faith."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws = wb.create_sheet("Confusion Matrix")
    title = Font(bold=True, size=14, color="1B5FD1")
    section = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1B5FD1")
    bold = Font(bold=True)
    small = Font(size=9, italic=True, color="666666")
    box = Border(*[Side(style="thin", color="999999")] * 4)
    green = PatternFill("solid", fgColor="D6F0DE")
    red = PatternFill("solid", fgColor="FBD9D9")
    grey = PatternFill("solid", fgColor="EDEDED")
    wrap = Alignment(wrap_text=True, vertical="top")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, width in zip("ABCDEF", (34, 26, 26, 30, 16, 16)):
        ws.column_dimensions[col].width = width

    def section_row(text: str) -> int:
        ws.append([])
        ws.append([text])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).font = section
        ws.cell(row=r, column=1).fill = section_fill
        return r

    ws.append(["Confusion matrix — and how to read it"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"].font = title

    ws.append(["Every field name is sorted into one of four boxes: did the pipeline produce it, and does the "
               "hand-written ground truth actually have it? The counts below are for FIELD NAMES across all "
               "53 documents."])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1).alignment = wrap
    ws.row_dimensions[2].height = 30

    # ---- the four buckets, with a real example of each
    section_row("STEP 1 — the four boxes, with a real example from your own data")
    ws.append(["Box", "Plain meaning", "Real example", "Why it lands there", "Count now", "Count before"])
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = grey
        ws.cell(row=ws.max_row, column=c).alignment = centre

    t, pt = current["totals"], (previous["totals"] if previous else {})
    tp_ex = examples.get("tp") or ("—", "—")
    fp_ex = examples.get("fp") or ("—", "—")
    fn_ex = examples.get("fn") or ("—", "—")
    rows = [
        ("TRUE POSITIVE (TP)", "Correct. Both the human and the pipeline recorded this field.",
         f'"{tp_ex[1]}"  ({tp_ex[0]})', "The pipeline used the same field name the human did.",
         t["key_tp"], pt.get("key_tp", "n/a"), green),
        ("FALSE POSITIVE (FP)", "Extra. The pipeline produced a field the ground truth has no such name for.",
         f'"{fp_ex[1]}"  ({fp_ex[0]})',
         "Often NOT an error of reading — usually the same value under a different name.",
         t["key_fp"], pt.get("key_fp", "n/a"), red),
        ("FALSE NEGATIVE (FN)", "Missed. The ground truth has this field; the pipeline did not produce that name.",
         f'"{fn_ex[1]}"  ({fn_ex[0]})',
         "Either genuinely not extracted, or extracted under a different name (then it also shows as an FP).",
         t["key_fn"], pt.get("key_fn", "n/a"), red),
        ("TRUE NEGATIVE (TN)", "Not possible to count here.", "—",
         "TN means 'correctly did not produce a field'. There is no fixed list of all possible field names, "
         "so this box cannot be filled — see the note at the bottom.",
         "n/a", "n/a", grey),
    ]
    for name, meaning, example, why, now, before, fill in rows:
        ws.append([name, meaning, example, why, now, before])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=1).fill = fill
        for c in range(1, 7):
            ws.cell(row=r, column=c).alignment = wrap
            ws.cell(row=r, column=c).border = box
        ws.row_dimensions[r].height = 42

    # ---- the grid
    section_row("STEP 2 — the same four numbers as a grid")
    ws.append(["", "Ground truth HAS this field", "Ground truth does NOT have it", "", "", ""])
    r = ws.max_row
    for c in (2, 3):
        ws.cell(row=r, column=c).font = bold
        ws.cell(row=r, column=c).fill = grey
        ws.cell(row=r, column=c).alignment = centre
        ws.cell(row=r, column=c).border = box

    for label, left, right, lfill, rfill in [
        ("Pipeline PRODUCED the field", f"TP = {t['key_tp']}", f"FP = {t['key_fp']}", green, red),
        ("Pipeline did NOT produce it", f"FN = {t['key_fn']}", "TN = n/a", red, grey),
    ]:
        ws.append([label, left, right])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=1).fill = grey
        ws.cell(row=r, column=2).fill = lfill
        ws.cell(row=r, column=3).fill = rfill
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).alignment = centre
        ws.row_dimensions[r].height = 26

    # ---- the arithmetic, spelled out
    section_row("STEP 3 — what those numbers give you (arithmetic shown, nothing hidden)")
    ws.append(["Measure", "Formula", "With your numbers", "In plain words", "Now", "Before"])
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = grey
        ws.cell(row=ws.max_row, column=c).alignment = centre

    cm, pm = current["metrics"], (previous["metrics"] if previous else {})
    calc = [
        ("Precision", "TP / (TP + FP)", f"{t['key_tp']} / ({t['key_tp']} + {t['key_fp']}) = {t['key_tp'] + t['key_fp']}",
         "When the pipeline names a field, how often is that a field the human also recorded.",
         _pct(cm["key_precision"]), _pct(pm.get("key_precision"))),
        ("Recall", "TP / (TP + FN)", f"{t['key_tp']} / ({t['key_tp']} + {t['key_fn']}) = {t['key_tp'] + t['key_fn']}",
         "Of the fields the human recorded, how many the pipeline also named.",
         _pct(cm["key_recall"]), _pct(pm.get("key_recall"))),
        ("F1 score", "2 × P × R / (P + R)", "the two above, balanced",
         "One number combining precision and recall. Use this for 'can it fill named form slots'.",
         _pct(cm["key_f1"]), _pct(pm.get("key_f1"))),
    ]
    for name, formula, subst, plain, now, before in calc:
        ws.append([name, formula, subst, plain, now, before])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        for c in range(1, 7):
            ws.cell(row=r, column=c).alignment = wrap
            ws.cell(row=r, column=c).border = box
        ws.row_dimensions[r].height = 34

    # ---- STRICT: name AND value must both be right ---------------------
    section_row("STEP 3B — STRICT scoring: the field name AND the value must both be right")
    ws.append(["Everything above scores NAMES only — a field counted as a TP even if its value was wrong. "
               "For a system whose job is to extract the field and its value together, that is too "
               "generous. Below, a prediction is correct ONLY when both halves match. This is the score "
               "that matches this product's goal."])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = wrap
    ws.row_dimensions[r].height = 44

    ws.append(["", "Ground truth pair", "No such ground-truth pair", "", "", ""])
    r = ws.max_row
    for c in (2, 3):
        ws.cell(row=r, column=c).font = bold
        ws.cell(row=r, column=c).fill = grey
        ws.cell(row=r, column=c).alignment = centre
        ws.cell(row=r, column=c).border = box
    for label, left, right, lfill, rfill in [
        ("Pipeline produced the pair", f"TP = {cm['strict_pair_tp']}", f"FP = {cm['strict_pair_fp']}", green, red),
        ("Pipeline did not produce it", f"FN = {cm['strict_pair_fn']}", "TN = n/a", red, grey),
    ]:
        ws.append([label, left, right])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=1).fill = grey
        ws.cell(row=r, column=2).fill = lfill
        ws.cell(row=r, column=3).fill = rfill
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).alignment = centre
        ws.row_dimensions[r].height = 26

    ws.append(["Measure", "Formula", "With your numbers", "In plain words", "Now", "Before"])
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = grey
        ws.cell(row=ws.max_row, column=c).alignment = centre
    stp, sfp, sfn = cm["strict_pair_tp"], cm["strict_pair_fp"], cm["strict_pair_fn"]
    for name, formula, subst, plain, now, before in [
        ("Strict precision", "TP / (TP + FP)", f"{stp} / ({stp} + {sfp}) = {stp + sfp}",
         "Of the pairs it output, how many were fully right — how much of the output is trustworthy as-is.",
         _pct(cm["strict_pair_precision"]), _pct(pm.get("strict_pair_precision"))),
        ("Strict recall", "TP / (TP + FN)", f"{stp} / ({stp} + {sfn}) = {stp + sfn}",
         "Of the pairs the human recorded, how many it got fully right — how much work it saves the user.",
         _pct(cm["strict_pair_recall"]), _pct(pm.get("strict_pair_recall"))),
        ("Strict F1", "2 x P x R / (P + R)", "the two above, balanced",
         "THE headline number for 'extract field and value'. Quote this one.",
         _pct(cm["strict_pair_f1"]), _pct(pm.get("strict_pair_f1"))),
    ]:
        ws.append([name, formula, subst, plain, now, before])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        for c in range(1, 7):
            ws.cell(row=r, column=c).alignment = wrap
            ws.cell(row=r, column=c).border = box
        ws.row_dimensions[r].height = 34

    ws.append(["Where the strict score loses points vs the name-only score above:", "", "", "", "", ""])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([f"  {cm['value_wrong_on_matched_keys']} pairs had the RIGHT field name but a WRONG value. "
               f"STEP 1-3 counted those as true positives; strict scoring counts each of them as one FP "
               f"and one FN, which is why strict F1 ({_pct(cm['strict_pair_f1'])}) sits below key F1 "
               f"({_pct(cm['key_f1'])})."])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = wrap
    ws.row_dimensions[r].height = 44

    # ---- the fairer, naming-independent view
    section_row("STEP 4 — the fairer view: ignore names, did it find the INFORMATION?")
    ws.append(["Boxes above compare NAMES. A correct value under a different name counts as BOTH an FP and an "
               "FN — two errors for one correct read. This view ignores names entirely."])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = wrap
    ws.row_dimensions[r].height = 30

    ws.append(["", "Found", "Not found", "Total", "Accuracy", ""])
    for c in range(1, 6):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = grey
        ws.cell(row=ws.max_row, column=c).alignment = centre
    for label, run in [("NOW", current)] + ([("BEFORE", previous)] if previous else []):
        tt = run["totals"]
        total = tt["gt_values"] + tt["gt_tests"]
        hit = tt["gt_values_found"] + tt["gt_tests_found"]
        ws.append([label, hit, total - hit, total, f"{hit / total:.1%}" if total else "n/a"])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=2).fill = green
        ws.cell(row=r, column=3).fill = red
        ws.cell(row=r, column=5).font = Font(bold=True, size=12, color="1A7F37")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).alignment = centre

    # ---- value-level
    section_row("STEP 5 — where both sides used the SAME name, was the value right?")
    ws.append(["", "Value matched", "Value differed", "Accuracy", "", ""])
    for c in range(1, 5):
        ws.cell(row=ws.max_row, column=c).font = bold
        ws.cell(row=ws.max_row, column=c).fill = grey
        ws.cell(row=ws.max_row, column=c).alignment = centre
    for label, run in [("NOW", current)] + ([("BEFORE", previous)] if previous else []):
        tt = run["totals"]
        matched = tt["value_correct"] + tt["value_wrong"]
        ws.append([label, tt["value_correct"], tt["value_wrong"],
                   f"{tt['value_correct'] / matched:.1%}" if matched else "n/a"])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=2).fill = green
        ws.cell(row=r, column=3).fill = red
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).alignment = centre

    # ---- the honest caveat
    section_row("IMPORTANT — why one box is empty")
    ws.append(["A textbook confusion matrix has four filled boxes and computes accuracy as (TP+TN)/everything. "
               "That cannot be done here. TN would mean 'a field the pipeline correctly did not produce', but "
               "field names are invented from each document's own labels — there is no fixed list of all "
               "possible names, so there is nothing to count. Putting 0 there would be wrong and would make "
               "any accuracy derived from it look far better than reality. The two honest headline numbers are "
               "therefore STRICT F1 (STEP 3B) when field name and value must both be right — the score for "
               "this product's goal — and COVERAGE (STEP 4) when the question is only whether the "
               "information was read at all."])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).alignment = wrap
    ws.cell(row=r, column=1).font = small
    ws.row_dimensions[r].height = 74


def build_workbook(current: dict, previous: dict | None, out_path: Path, examples: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1B5FD1")
    bold = Font(bold=True)
    note = Font(italic=True, size=9, color="666666")
    good = PatternFill("solid", fgColor="E7F6EC")
    bad = PatternFill("solid", fgColor="FDECEC")

    wb = Workbook()

    def style_header(ws, row=1, width=None):
        for cell in ws[row]:
            if cell.value is not None:
                cell.font, cell.fill = head, head_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, w in enumerate(width or [], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------------------------------------------------------- Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Previous version", "Current version", "Change", "What it means"])
    style_header(ws, width=[36, 17, 17, 12, 78])

    cm, pm = current["metrics"], (previous["metrics"] if previous else {})
    spec = [
        ("STRICT ACCURACY (F1) — field + value", "strict_pair_f1",
         "THE SCORE FOR THIS PRODUCT'S GOAL. A prediction counts as correct only when the field NAME and "
         "the VALUE are both right. Right name with a wrong value counts as an error, not a hit. This is "
         "entity-level scoring, the standard for open-set key-value extraction."),
        ("  · strict precision", "strict_pair_precision",
         "Of the field+value pairs the pipeline produced, how many were exactly right. This is how much of "
         "the output a reviewer can trust as-is."),
        ("  · strict recall", "strict_pair_recall",
         "Of the field+value pairs the ground truth records, how many the pipeline reproduced exactly. This "
         "is how much of the work the pipeline actually does for the user."),
        ("OVERALL ACCURACY (coverage)", "overall_accuracy",
         "Share of EVERYTHING the ground truth records (fields + test rows) that the pipeline extracted, "
         "regardless of what it named the field. The single headline number."),
        ("  · document-field coverage", "field_coverage",
         "Of the ground truth's header/metadata field values, how many were found under any key."),
        ("  · test-result coverage", "test_result_coverage",
         "Of the ground truth's test-table results, how many appear in the extraction (open extraction "
         "picks table contents up as flat fields)."),
        ("KEY PRECISION", "key_precision",
         "Of the field NAMES the pipeline produced, how many exist in the ground truth. Low = it invents "
         "or splits names, not necessarily that it read badly."),
        ("KEY RECALL", "key_recall", "Of the ground truth's field NAMES, how many the pipeline also used."),
        ("KEY F1", "key_f1",
         "Harmonic mean of key precision/recall — the strict, name-sensitive score. Use this one when the "
         "question is 'can it fill named form slots'."),
        ("VALUE ACCURACY (on matched keys)", "value_accuracy_on_matched_keys",
         "Where both sides used the same field name, how often the VALUE also matched."),
        ("Documents scored", "documents", "Total documents in the ground-truth set."),
        ("Documents failed", "documents_failed", "Extraction raised an error and produced nothing."),
    ]
    for label, key, meaning in spec:
        cur, prev = cm.get(key), pm.get(key)
        is_count = key in ("documents", "documents_failed")
        change = ""
        if not is_count and cur is not None and prev is not None:
            change = f"{(cur - prev) * 100:+.1f} pts"
        ws.append([
            label,
            (prev if is_count else _pct(prev)) if previous else "n/a",
            cur if is_count else _pct(cur),
            change,
            meaning,
        ])
        row = ws.max_row
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        if label.startswith(("STRICT", "OVERALL", "KEY", "VALUE")):
            ws.cell(row=row, column=1).font = bold
        if change.startswith("+"):
            ws.cell(row=row, column=4).fill = good
        elif change.startswith("-"):
            ws.cell(row=row, column=4).fill = bad

    ws.append([])
    ws.append(["NOTE", "", "", "",
               "TRUE NEGATIVES are undefined here: open-set extraction has no fixed list of possible field "
               "names, so there is no such thing as a field correctly NOT extracted. Textbook accuracy "
               "(TP+TN)/total therefore cannot be computed; OVERALL ACCURACY above is coverage instead."])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.cell(row=ws.max_row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    # ------------------------------------------------------- Confusion matrix
    _confusion_sheet(wb, current, previous, examples)

    # --------------------------------------------------------- Per-document
    ws = wb.create_sheet("Per-document")
    cols = ["document", "status", "gt_fields", "pred_fields", "key_tp", "key_fp", "key_fn",
            "value_correct", "value_wrong", "gt_values", "gt_values_found", "gt_tests", "gt_tests_found"]
    header = ["Document", "Status", "GT fields", "Pred fields", "TP", "FP", "FN",
              "Value ok", "Value wrong", "GT values", "Found", "GT tests", "Tests found", "Coverage %"]
    if previous:
        header += ["Prev coverage %", "Change pts"]
    ws.append(header)
    style_header(ws, width=[46, 9] + [11] * 12 + [16, 12])

    prev_rows = {r["document"]: r for r in previous["rows"]} if previous else {}
    for r in current["rows"]:
        total = r["gt_values"] + r["gt_tests"]
        hit = r["gt_values_found"] + r["gt_tests_found"]
        cov = hit / total if total else None
        line = [r[c] for c in cols] + [round(cov * 100, 1) if cov is not None else "n/a"]
        if previous:
            pr = prev_rows.get(r["document"])
            ptotal = (pr["gt_values"] + pr["gt_tests"]) if pr else 0
            phit = (pr["gt_values_found"] + pr["gt_tests_found"]) if pr else 0
            pcov = phit / ptotal if ptotal else None
            line += [
                round(pcov * 100, 1) if pcov is not None else "n/a",
                round((cov - pcov) * 100, 1) if (cov is not None and pcov is not None) else "n/a",
            ]
        ws.append(line)
        if r["status"] == "failed":
            for c in range(1, len(line) + 1):
                ws.cell(row=ws.max_row, column=c).fill = bad

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--labelled", default=os.path.join(DEFAULT_ROOT, "labelled_dataset"))
    parser.add_argument("--current", default=os.path.join(DEFAULT_ROOT, "predictions_ui_pipeline", "predictions"),
                        help="Current run's predictions/ folder.")
    parser.add_argument("--previous", default=r"D:\pip_tmp\baseline_preds",
                        help="Previous run's predictions/ folder (omit or point at a missing path to skip).")
    parser.add_argument("--out", default=os.path.join(DEFAULT_ROOT, "benchmark_report.xlsx"))
    args = parser.parse_args()

    labelled_dir = Path(args.labelled)
    if not labelled_dir.is_dir():
        print(f"Labelled dataset not found: {labelled_dir}", file=sys.stderr)
        return 1

    current = evaluate(labelled_dir, Path(args.current))
    prev_dir = Path(args.previous)
    previous = evaluate(labelled_dir, prev_dir) if prev_dir.is_dir() else None
    if previous is None:
        print(f"(no previous run at {prev_dir} — reporting current only)")

    out_path = Path(args.out)
    examples = pick_examples(labelled_dir, Path(args.current))
    build_workbook(current, previous, out_path, examples)

    print(f"Wrote {out_path}")
    print()
    print(f"{'metric':38} {'previous':>12} {'current':>12}")
    for k in ("strict_pair_f1", "strict_pair_precision", "strict_pair_recall",
              "overall_accuracy", "field_coverage", "test_result_coverage",
              "key_precision", "key_recall", "key_f1", "value_accuracy_on_matched_keys"):
        prev = _pct(previous["metrics"][k]) if previous else "n/a"
        print(f"{k:38} {prev:>12} {_pct(current['metrics'][k]):>12}")
    print(f"{'documents / failed':38} "
          f"{(str(previous['metrics']['documents']) + ' / ' + str(previous['metrics']['documents_failed'])) if previous else 'n/a':>12} "
          f"{str(current['metrics']['documents']) + ' / ' + str(current['metrics']['documents_failed']):>12}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
