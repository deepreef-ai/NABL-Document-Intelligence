#!/usr/bin/env python3
"""Step 2 of the 2-script pipeline: for every document already labeled in
labelled_dataset/ (see create_ground_truth.py), ask an LLM — Amazon Nova on
Bedrock by default, see app/llm/providers.py's NovaProvider — to extract the
same document fresh, cache the result to predictions/<name>.json (resumable:
already-cached documents are skipped unless --force), then score every
cached prediction against its ground truth and write one summary.

No separate "normalize" stage or cached text folder: text/image extraction
happens per-document, in-process. A born-digital page's real text comes
straight from app/documents/pdf_utils.py (pure PyMuPDF). A page with no text
layer (scanned PDF or a raw image file) is rasterized and run through local
OCR (app/documents/local_ocr.py's extract_english — RapidOCR's bundled
English model, not deepreef-ocr's Devanagari model, which garbles English
digits/table layout badly enough to misattribute whole rows — see this
project's OCR comparison notes) plus text_repair's glued-word fix.

The OCR/PyMuPDF text always goes to Nova; the ORIGINAL image goes too
whenever exactly one image represents the whole document (a raw image file,
or a single-page PDF) — MEASURED: text-only extraction cannot tell a blank/
redacted box from a value that legitimately looks like the next field's own
label, since OCR emits no signal at all for "there was a gap here" (see
001_Lab-report.png: patient_name/patient_id/accession_no/... all came back
as their own label text, e.g. "patient_name": "PATIENT NAME" — a prompt-only
fix did not resolve it). A multi-page PDF still gets text-only: its pages
are concatenated into one combined-document call, and there's no single
image that represents "the whole document" to attach to that call — fixing
that would mean extracting per-page like the production pipeline
(documents/pipeline.py) does, which is a bigger restructuring than this
script currently does.

Usage:
    python scripts/generate_predictions_and_score.py
    python scripts/generate_predictions_and_score.py --force
    python scripts/generate_predictions_and_score.py --score-only
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # backend/ on sys.path

from app.benchmark.accumulator import MetricAccumulator  # noqa: E402
from app.benchmark.compare import build_test_comparison_rows, compare_fields, compare_tests, field_comparison_rows  # noqa: E402
from app.config import get_settings  # noqa: E402
from app.dataset_normalization.text_repair import repair_glued_words  # noqa: E402
from app.documents import local_ocr, pdf_utils  # noqa: E402
from app.documents.app import extract_lab_report  # noqa: E402
from app.llm.chain import LlmChain  # noqa: E402
from app.llm.factory import _build_chain  # noqa: E402

DEFAULT_ROOT = r"G:\Shared drives\Product & Engineering\Projects\NABL Document Intelligence"
# Longest provider cooldown this batch will sit and wait through. Above
# this, recording the failure and moving on beats stalling the whole run
# (LlmChain caps its own backoff at 300s).
_MAX_COOLDOWN_WAIT_SECONDS = 120.0

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff"}
_IMAGE_MEDIA_TYPES = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".tif": "image/tiff", ".tiff": "image/tiff"}


def _find_source_file(dataset_dir: Path, filename: str) -> Path | None:
    candidate = dataset_dir / filename
    if candidate.is_file():
        return candidate
    # dataset/ is organized into subfolders (food/, medical/, ...) — the
    # ground truth only records the bare filename, so fall back to a
    # recursive search.
    matches = [p for p in dataset_dir.rglob(filename) if p.is_file()]
    return matches[0] if matches else None


def _ocr_image_to_text(image_bytes: bytes) -> str:
    result = local_ocr.extract_english(image_bytes)
    return repair_glued_words(result.text)


def _build_message_content(source_path: Path) -> tuple[str, bytes | None, str | None]:
    """Returns (user_text, image_bytes, image_media_type). An image is
    attached only when exactly one image legitimately represents the WHOLE
    document — a raw image file, or a single-page PDF (rasterized here
    regardless of whether it already has a text layer, purely to get pixels
    to send alongside the text). A multi-page PDF's pages are concatenated
    into one combined-document text blob for a single extraction call, and
    there's no one image that represents "the whole document" to attach to
    that call, so it stays text-only — see this module's docstring."""
    suffix = source_path.suffix.lower()
    data = source_path.read_bytes()

    if suffix in IMAGE_EXTENSIONS:
        return _ocr_image_to_text(data), data, _IMAGE_MEDIA_TYPES[suffix]

    if suffix != ".pdf":
        raise ValueError(f"unsupported file type: {suffix}")

    pages = pdf_utils.extract_text_and_boxes(data)
    count = pdf_utils.page_count(data)
    text_blocks: list[str] = []
    for i in range(count):
        page = next((p for p in pages if p.page_number == i), None)
        text = page.text if page else ""
        if len(text.strip()) < 20:
            png_bytes = pdf_utils.rasterize_page(data, i)
            text = _ocr_image_to_text(png_bytes)
        text_blocks.append(f"--- Page {i + 1} ---\n{text}")
    user_text = "\n\n".join(text_blocks)

    if count == 1:
        return user_text, pdf_utils.rasterize_page(data, 0), "image/png"
    return user_text, None, None


def _predict_one(chain: LlmChain, source_path: Path) -> dict:
    user_text, image_bytes, image_media_type = _build_message_content(source_path)
    result = extract_lab_report(chain, user_text, image=image_bytes, image_media_type=image_media_type)
    return {**result, "pipeline_error": None}


def generate_predictions(labelled_dir: Path, dataset_dir: Path, predictions_dir: Path, chain: LlmChain, force: bool = False) -> dict:
    predictions_dir.mkdir(parents=True, exist_ok=True)
    stats = {"total": 0, "predicted": 0, "skipped": 0, "failed": 0, "failures": []}

    for label_path in sorted(labelled_dir.glob("*/*.json")):
        stem = label_path.stem
        out_path = predictions_dir / f"{stem}.json"
        stats["total"] += 1
        if out_path.exists() and not force:
            # A cached FAILURE is not "already done" — a timeout or a
            # transient chain cooldown is worth retrying automatically on
            # the next run; only a genuine success is skipped. --force
            # still means "redo everything, successes included".
            cached = json.loads(out_path.read_text(encoding="utf-8"))
            if cached.get("pipeline_error") is None:
                stats["skipped"] += 1
                continue

        ground_truth = json.loads(label_path.read_text(encoding="utf-8"))
        source_path = _find_source_file(dataset_dir, ground_truth["original_filename"])
        if source_path is None:
            prediction = {"fields": {}, "tests": [], "pipeline_error": f"source file not found: {ground_truth['original_filename']}"}
        else:
            # This is a BATCH job, so waiting out a provider cooldown is right —
            # the chain itself fails fast because an interactive request must
            # not block, but here that means abandoning work. MEASURED
            # 2026-09-04: one rate-limit cascade cost 37 of 51 documents, each
            # failing instantly with "cooling down for Ns more" without a
            # single API attempt. See LlmChain.seconds_until_available.
            wait = chain.seconds_until_available()
            if wait is not None and wait != float("inf") and wait <= _MAX_COOLDOWN_WAIT_SECONDS:
                print(f"  (all providers cooling down, waiting {wait:.0f}s before {stem})", flush=True)
                time.sleep(wait + 1)
            try:
                prediction = _predict_one(chain, source_path)
            except Exception as exc:  # noqa: BLE001 — one bad document must never stop the batch
                prediction = {"fields": {}, "tests": [], "pipeline_error": f"{type(exc).__name__}: {exc}"}

        out_path.write_text(json.dumps(prediction, indent=2, ensure_ascii=False), encoding="utf-8")
        if prediction["pipeline_error"]:
            stats["failed"] += 1
            stats["failures"].append((stem, prediction["pipeline_error"]))
        else:
            stats["predicted"] += 1

    return stats


def score(labelled_dir: Path, predictions_dir: Path) -> tuple[dict, list[dict], list[dict], list[dict]]:
    overall = MetricAccumulator()
    rows: list[dict] = []
    all_field_rows: list[dict] = []
    all_test_rows: list[dict] = []

    for label_path in sorted(labelled_dir.glob("*/*.json")):
        stem = label_path.stem
        ground_truth = json.loads(label_path.read_text(encoding="utf-8"))
        pred_path = predictions_dir / f"{stem}.json"

        if pred_path.exists():
            prediction = json.loads(pred_path.read_text(encoding="utf-8"))
        else:
            prediction = {"fields": {}, "tests": [], "pipeline_error": "no cached prediction found"}

        extraction_ok = prediction.get("pipeline_error") is None
        gt_fields, pred_fields = ground_truth.get("fields", {}), prediction.get("fields", {})
        gt_tests, pred_tests = ground_truth.get("tests", []), prediction.get("tests", [])
        field_failures, field_counters = compare_fields(stem, gt_fields, pred_fields)
        test_failures, test_counters = compare_tests(stem, gt_tests, pred_tests)
        exact_match = extraction_ok and not field_failures and not test_failures

        doc_acc = MetricAccumulator()
        doc_acc.add(domain_match=True, extraction_ok=extraction_ok, exact_match=exact_match,
                    field_counters=field_counters, test_counters=test_counters)
        overall.add(domain_match=True, extraction_ok=extraction_ok, exact_match=exact_match,
                    field_counters=field_counters, test_counters=test_counters)

        rows.append({
            "document": stem,
            "original_filename": ground_truth.get("original_filename"),
            "extraction_status": "ok" if extraction_ok else "failed",
            "pipeline_error": prediction.get("pipeline_error"),
            **doc_acc.finalize(),
        })
        all_field_rows.extend(field_comparison_rows(stem, gt_fields, pred_fields, prediction.get("field_verified")))
        all_test_rows.extend(build_test_comparison_rows(stem, gt_tests, pred_tests))

    return overall.finalize(), rows, all_field_rows, all_test_rows


def write_results(summary: dict, rows: list[dict], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")

    import csv
    if rows:
        with (output_dir / "per_document.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)


# Status -> fill color for the comparison sheets, so a reviewer can scan a
# whole sheet by eye instead of reading the "status" column cell by cell.
_STATUS_COLORS = {
    "correct": "C6EFCE", "missing": "FFEB9C", "wrong_value": "FFC7CE",
    "wrong_key": "FFC7CE", "wrong_unit": "FFC7CE", "extra": "E4C7FF",
}


def _cell_value(value):
    # A ground-truth/predicted "value" is occasionally a list (e.g. a
    # multi-line field the labeler or the model captured as several
    # strings) or a dict — openpyxl only accepts Excel-native scalar types,
    # so anything else is flattened to its JSON text rather than crashing
    # the whole export over one odd cell.
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _write_sheet(ws, rows: list[dict], status_key: str = "status") -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not rows:
        ws.append(["(no rows)"])
        return

    columns = list(rows[0].keys())
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    status_col = columns.index(status_key) + 1 if status_key in columns else None
    for row in rows:
        ws.append([_cell_value(row.get(c)) for c in columns])
        if status_col:
            color = _STATUS_COLORS.get(row.get(status_key))
            if color:
                ws.cell(ws.max_row, status_col).fill = PatternFill("solid", fgColor=color)

    for i, column in enumerate(columns, start=1):
        width = max(len(str(column)), *(len(str(r.get(column, ""))) for r in rows[:200]))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)


def write_excel_workbook(
    summary: dict, rows: list[dict], field_rows: list[dict], test_rows: list[dict], output_dir: Path
) -> Path:
    """A human-reviewable comparison workbook alongside summary.json/
    per_document.csv — one row per ground-truth field/test (correct or
    not), built from the exact same compare_fields/compare_tests matching
    the scored metrics use (see field_comparison_rows/
    build_test_comparison_rows in app/benchmark/compare.py), so this can
    never disagree with the printed accuracy numbers.

    Returns the path actually written. Reviewing comparison.xlsx in Excel
    holds a Windows write lock on it, and this runs LAST — after every
    (paid, slow) LLM call — so a lock here used to throw away a whole
    completed run's reporting. Falls back to a timestamped filename
    instead of failing."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["metric", "value"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    for key, value in summary.items():
        summary_ws.append([key, value])
    summary_ws.append([])
    summary_ws.append(["--- per document ---"])
    if rows:
        columns = list(rows[0].keys())
        summary_ws.append(columns)
        for cell in summary_ws[summary_ws.max_row]:
            cell.font = Font(bold=True)
        for row in rows:
            summary_ws.append([row.get(c) for c in columns])
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 20

    _write_sheet(wb.create_sheet("Fields"), field_rows)
    _write_sheet(wb.create_sheet("Tests"), test_rows)

    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / "comparison.xlsx"
    try:
        wb.save(target)
        return target
    except PermissionError:
        from datetime import datetime
        fallback = output_dir / f"comparison-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
        wb.save(fallback)
        print(f"\nNOTE: {target.name} is open in another program (locked) — wrote {fallback.name} instead.", file=sys.stderr)
        return fallback


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--labelled", default=os.path.join(DEFAULT_ROOT, "labelled_dataset"), help="Ground truth directory (create_ground_truth.py's output).")
    parser.add_argument("--dataset", default=os.path.join(DEFAULT_ROOT, "dataset"), help="Raw source files directory.")
    parser.add_argument("--output", default=None, help="Output directory for predictions/ and summary. Defaults to a 'predictions' folder next to --labelled.")
    parser.add_argument("--force", action="store_true", help="Re-run the LLM call even for already-predicted documents.")
    parser.add_argument("--predict-only", action="store_true")
    parser.add_argument("--score-only", action="store_true")
    parser.add_argument(
        "--providers", default=None,
        help="Override LLM_PROVIDER_ORDER for this run, e.g. 'gemini' or 'nova,gemini'. "
             "Useful to pin the benchmark to one provider so results stay comparable.",
    )
    parser.add_argument(
        "--timeout", type=float, default=180.0,
        help="Per-call timeout in seconds (default: 180). The app's usual LLM_TIMEOUT_SECONDS (30s, tuned for "
             "short chat-style calls elsewhere) is too short here — a document with a genuinely large results "
             "table (see high-protein-paneer.pdf's 269 test rows) needs real headroom to both generate and "
             "transfer a long JSON reply.",
    )
    args = parser.parse_args()

    labelled_dir = Path(args.labelled)
    if not labelled_dir.is_dir():
        print(f"Labelled dataset not found: {labelled_dir}", file=sys.stderr)
        return 1

    dataset_dir = Path(args.dataset)
    output_dir = Path(args.output) if args.output else labelled_dir.parent / "predictions"
    predictions_dir = output_dir / "predictions"

    settings = get_settings()
    # Honour LLM_PROVIDER_ORDER rather than hardcoding Nova, so a Bedrock
    # outage can be ridden out on the fallback provider (MEASURED 2026-09-04:
    # Bedrock returned "Operation not allowed" account-wide for a full day,
    # which would otherwise have stopped the benchmark entirely).
    #
    # --timeout overrides llm_timeout_seconds for this run: the app's 30s
    # default is tuned for short chat turns, whereas one document here is a
    # full-page image plus its OCR text plus a ~1.5k-token system prompt, and
    # a large results table takes far longer than 30s to generate.
    settings = settings.model_copy(update={"llm_timeout_seconds": args.timeout})
    chain = _build_chain(settings, args.providers or settings.llm_provider_order)
    if not chain.providers:
        print(
            f"No LLM provider configured for order {args.providers or settings.llm_provider_order!r} — "
            "set NOVA_MODEL and/or GEMINI_API_KEY in backend/.env (see .env.example).",
            file=sys.stderr,
        )
        return 1
    print(f"LLM providers:    {', '.join(p.name for p in chain.providers)} (timeout {args.timeout:.0f}s)")

    print(f"Labelled dataset: {labelled_dir}")
    print(f"Dataset (raw):    {dataset_dir}")
    print(f"Output:           {output_dir}")

    if not args.score_only:
        stats = generate_predictions(labelled_dir, dataset_dir, predictions_dir, chain, force=args.force)
        print()
        print("=== Prediction Summary ===")
        print(f"Total documents: {stats['total']}")
        print(f"Predicted (new): {stats['predicted']}")
        print(f"Skipped (cached): {stats['skipped']}")
        print(f"Failed:          {stats['failed']}")
        for name, reason in stats["failures"]:
            print(f"  {name}: {reason}")

    if not args.predict_only:
        summary, rows, field_rows, test_rows = score(labelled_dir, predictions_dir)
        write_results(summary, rows, output_dir)
        workbook_path = write_excel_workbook(summary, rows, field_rows, test_rows, output_dir)
        print()
        print("=== Score Summary ===")
        for key, value in summary.items():
            print(f"{key}: {value}")
        print(f"\nComparison workbook: {workbook_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
